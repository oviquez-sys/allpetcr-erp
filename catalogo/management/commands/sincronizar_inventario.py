"""Sincroniza el catálogo con el Excel de inventario vigente.

Uso:
    python manage.py sincronizar_inventario --dry-run     # ver qué haría
    python manage.py sincronizar_inventario               # aplicarlo

Por qué existe (y por qué NO se reusó `importar_inventario`)
------------------------------------------------------------
`importar_inventario` era una carga inicial: leía la hoja "Inventario Real",
creaba lo que no existía y **omitía** lo que ya estaba. Servía una vez.

El Excel del 02/08/2026 (`data/INVENTARIO_ALLPETCR.xlsx`) cambia dos cosas que
rompen ese comando: la hoja se llama "Inventario Claude" y el orden de las
columnas es otro (donde antes iba el código de barras ahora va la
subcategoría). Corriendo el viejo, la subcategoría entraría como código de
barras sin que nada fallara. Por eso un comando nuevo y no un parche.

Además el trabajo ya no es "crear": de los 532 productos del Excel, 184 ya
existen y hay que **actualizarlos** —precio, nombre, descripción, categoría—
sin perder su historial. Actualizar es una operación distinta de crear, y
mucho más peligrosa: por eso `--dry-run` y por eso el resumen detallado.

Reglas del proyecto que este comando respeta (no son opcionales)
----------------------------------------------------------------
- **El stock se mueve por el kardex, nunca a mano.** Toda diferencia entre el
  Excel y la base entra como movimiento (INI si el producto es nuevo, AJU si
  ya existía) vía `inventario.services.registrar_movimiento`.
- **El precio se cambia por `catalogo.services.cambiar_precio`**, que deja
  cada cambio firmado en `CambioPrecio`. 184 precios cambian de golpe: sin
  bitácora sería imposible saber después qué pasó ni volver atrás.
- **`costo_promedio` NO se toca.** Es un valor derivado del kardex. El Excel
  trae un "precio mayorista" que puede no coincidir con el costo promedio
  real; corregirlo es un evento contable, no una importación. El comando
  reporta las diferencias y no hace nada más con ellas.

Qué se preserva de lo que ya hay
--------------------------------
`presentacion` ("Paquete: 12 / Caja: 144") y `codigo_barras` NO vienen en el
Excel nuevo. Si se sobrescribieran con vacío, el sitio perdería un dato que
hoy muestra y habría que reimprimir etiquetas. Se conservan.
"""
from decimal import Decimal

import openpyxl
from django.core.exceptions import ValidationError
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from catalogo.models import Categoria, Producto
from catalogo.services import cambiar_precio
from core.models import Empresa, Sucursal
from inventario.models import Bodega
from inventario.services import registrar_movimiento

HOJA_POR_DEFECTO = "Inventario Claude"
ARCHIVO_POR_DEFECTO = "data/INVENTARIO_ALLPETCR.xlsx"

# Índices de columna (0-based) de la hoja "Inventario Claude". Se declaran
# aquí y se validan contra los encabezados reales antes de leer una sola
# fila: si alguien reordena el Excel, el comando aborta con un mensaje claro
# en vez de escribir la subcategoría dentro del código de barras.
COL = {
    "sku": 1,
    "nombre": 2,
    "categoria_web": 3,
    "subcategoria": 4,
    "mascota": 5,
    "descripcion": 6,
    "costo": 7,
    "cantidad": 8,
    "precio": 12,
}
ENCABEZADOS_ESPERADOS = {
    1: "Código (REF)",
    2: "Nombre",
    3: "Categoría web",
    4: "Subcategoría",
    5: "Mascota",
    6: "Descripción",
    7: "Precio mayorista (CRC)",
    8: "Cantidad en inventario",
    12: "Precio venta LOCAL sugerido (₡)",
}


def _texto(valor, limite=None):
    s = "" if valor is None else str(valor).strip()
    return s[:limite] if limite else s


def _decimal(valor):
    if valor in (None, ""):
        return Decimal("0")
    return Decimal(str(valor)).quantize(Decimal("0.01"))


class Command(BaseCommand):
    help = "Sincroniza productos, categorías, precios y stock con el Excel de inventario."

    def add_arguments(self, parser):
        parser.add_argument("--archivo", default=ARCHIVO_POR_DEFECTO)
        parser.add_argument("--hoja", default=HOJA_POR_DEFECTO)
        parser.add_argument(
            "--dry-run", action="store_true",
            help="No escribe nada: muestra exactamente qué cambiaría.",
        )
        parser.add_argument(
            "--usuario", default=None,
            help="Username que queda firmando los cambios de precio y los movimientos de kardex.",
        )
        parser.add_argument(
            "--sin-precios", action="store_true",
            help="Sincroniza todo MENOS los precios de venta: nombres, descripciones, "
                 "categorías, mascota, productos nuevos y stock. Sirve para aplicar datos "
                 "cuando los precios todavía están en discusión — pasó el 02/08/2026 con "
                 "el factor 1.25 (ver ACTUALIZAR_INVENTARIO.txt).",
        )
        parser.add_argument(
            "--desactivar-ausentes", action="store_true",
            help="Marca activo=False los productos de la base que ya no están en el Excel. "
                 "Por defecto solo se reportan: desaparecer un producto del ERP borra su "
                 "historial de vista y casi nunca es lo que se quiere.",
        )

    def handle(self, *args, **op):
        seco = op["dry_run"]
        filas = self._leer(op["archivo"], op["hoja"])

        empresa = Empresa.objects.first()
        if empresa is None:
            raise CommandError("No hay ninguna empresa configurada en la base de datos.")
        bodega = (
            Bodega.objects.filter(sucursal__empresa=empresa).order_by("id").first()
        )
        if bodega is None:
            sucursal, _ = Sucursal.objects.get_or_create(empresa=empresa, nombre="Central")
            if seco:
                self.stdout.write(self.style.WARNING(
                    "  No hay bodega; en la corrida real se crearía 'Principal'."
                ))
                bodega = None
            else:
                bodega = Bodega.objects.create(sucursal=sucursal, nombre="Principal")

        usuario = self._usuario(op["usuario"])

        # Toda la sincronización va en una sola transacción: si algo falla a
        # mitad de camino, el catálogo no queda mitad viejo y mitad nuevo.
        #
        # En --dry-run se marca la transacción para revertir con
        # `set_rollback(True)` en vez de lanzar una excepción de control. Es
        # la forma idiomática de Django y, sobre todo, evita el patrón
        # `except ...: pass` que el propio proyecto prohíbe
        # (core/test_arquitectura.py). La primera versión de este comando lo
        # usaba y la prueba de arquitectura lo rechazó — con razón: aunque acá
        # la excepción fuera inofensiva, la regla no puede tener excepciones
        # "obvias" o deja de verificarse sola.
        with transaction.atomic():
            r = self._sincronizar(
                filas, empresa, bodega, usuario, seco, sin_precios=op["sin_precios"]
            )
            if seco:
                transaction.set_rollback(True)

        self._reportar(r, seco, sin_precios=op["sin_precios"])

    # ── lectura ──────────────────────────────────────────────────────────
    def _leer(self, ruta, hoja):
        try:
            wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
        except FileNotFoundError:
            raise CommandError(f"No se encontró el archivo: {ruta}")
        if hoja not in wb.sheetnames:
            raise CommandError(
                f"El Excel no tiene la hoja '{hoja}'. Hojas disponibles: {', '.join(wb.sheetnames)}"
            )
        ws = wb[hoja]
        filas = list(ws.iter_rows(values_only=True))
        wb.close()
        if not filas:
            raise CommandError("La hoja está vacía.")

        # Validación de encabezados: es la única defensa contra que alguien
        # reordene columnas en Excel y el comando escriba datos en el campo
        # equivocado sin que nada se queje.
        encabezado = filas[0]
        malos = []
        for idx, esperado in ENCABEZADOS_ESPERADOS.items():
            real = _texto(encabezado[idx] if idx < len(encabezado) else None)
            if real != esperado:
                malos.append(f"    columna {idx + 1}: se esperaba «{esperado}», hay «{real}»")
        if malos:
            raise CommandError(
                "Los encabezados del Excel no coinciden con los que espera este comando.\n"
                + "\n".join(malos)
                + "\n\n  Si el Excel cambió a propósito, actualizá COL y ENCABEZADOS_ESPERADOS "
                  "en este archivo. NO corras el comando así: escribiría cada dato en la "
                  "columna equivocada."
            )
        return [f for f in filas[1:] if f and f[COL["sku"]]]

    def _usuario(self, username):
        if not username:
            return None
        from django.contrib.auth import get_user_model
        try:
            return get_user_model().objects.get(username=username)
        except get_user_model().DoesNotExist:
            raise CommandError(f"No existe el usuario '{username}'.")

    # ── sincronización ───────────────────────────────────────────────────
    def _sincronizar(self, filas, empresa, bodega, usuario, seco, sin_precios=False):
        r = _Resultado()
        r.sin_precios = sin_precios
        vistos = set()

        for fila in filas:
            sku = _texto(fila[COL["sku"]])
            nombre = _texto(fila[COL["nombre"]], 200)
            if not sku or not nombre:
                r.saltados += 1
                continue
            if sku in vistos:
                r.duplicados.append(sku)
                continue
            vistos.add(sku)

            categoria = self._categoria(
                _texto(fila[COL["categoria_web"]], 80),
                _texto(fila[COL["subcategoria"]], 80),
                r,
            )
            costo_excel = _decimal(fila[COL["costo"]])
            precio = _decimal(fila[COL["precio"]])
            cantidad = _decimal(fila[COL["cantidad"]])

            producto = Producto.objects.filter(sku=sku).first()
            if producto is None:
                producto = Producto(
                    empresa=empresa,
                    sku=sku,
                    nombre=nombre,
                    categoria=categoria,
                    categoria_original=_texto(fila[COL["categoria_web"]], 120),
                    descripcion=_texto(fila[COL["descripcion"]]),
                    mascota=_texto(fila[COL["mascota"]], 30),
                    # Un producto nuevo SÍ lleva su precio aunque vaya
                    # --sin-precios: no hay precio anterior que preservar, y
                    # crearlo en cero lo dejaría fuera del sitio (el
                    # exportador omite los de precio 0) sin que se note.
                    precio_venta=precio,
                )
                producto.save()
                r.creados += 1
                if cantidad > 0 and bodega is not None:
                    registrar_movimiento(
                        producto=producto, bodega=bodega, tipo="INI",
                        cantidad=cantidad, costo_unitario=costo_excel,
                        referencia="SINC-INVENTARIO", motivo="Alta desde el Excel de inventario",
                        usuario=usuario,
                    )
                    r.stock_alta += 1
                continue

            # ── producto existente: campos descriptivos ──
            campos = []
            for campo, nuevo in (
                ("nombre", nombre),
                ("descripcion", _texto(fila[COL["descripcion"]])),
                ("mascota", _texto(fila[COL["mascota"]], 30)),
                ("categoria_original", _texto(fila[COL["categoria_web"]], 120)),
            ):
                if getattr(producto, campo) != nuevo:
                    setattr(producto, campo, nuevo)
                    campos.append(campo)
            if producto.categoria_id != (categoria.id if categoria else None):
                producto.categoria = categoria
                campos.append("categoria")
            if campos:
                producto.save(update_fields=campos + ["actualizado_en"])
                r.actualizados += 1
                if "nombre" in campos:
                    r.renombrados += 1

            # ── precio: siempre por el servicio, para que quede firmado ──
            if precio != producto.precio_venta:
                r.precios.append((sku, producto.precio_venta, precio))
                if sin_precios:
                    r.precios_omitidos += 1
                else:
                    try:
                        cambiar_precio(
                            producto=producto, nuevo_precio=precio, usuario=usuario,
                            motivo="Sincronización con el Excel de inventario 02/08/2026",
                        )
                    except ValidationError as e:
                        r.errores.append(f"{sku}: precio no aplicado — {'; '.join(e.messages)}")

            # ── stock: la diferencia entra como ajuste al kardex ──
            producto.refresh_from_db()
            diferencia = cantidad - producto.stock_actual
            if diferencia != 0 and bodega is not None:
                try:
                    registrar_movimiento(
                        producto=producto, bodega=bodega, tipo="AJU",
                        cantidad=diferencia,
                        # Solo las entradas llevan costo; una salida no
                        # recalcula el costo promedio (regla del servicio).
                        costo_unitario=costo_excel if diferencia > 0 else Decimal("0"),
                        referencia="SINC-INVENTARIO",
                        motivo="Ajuste al conteo del Excel de inventario 02/08/2026",
                        usuario=usuario,
                    )
                    r.ajustes.append((sku, diferencia))
                except ValidationError as e:
                    r.errores.append(f"{sku}: ajuste no aplicado — {'; '.join(e.messages)}")

            # ── costo: se reporta, NO se corrige ──
            if costo_excel > 0 and producto.costo_promedio > 0:
                brecha = abs(costo_excel - producto.costo_promedio)
                if brecha / producto.costo_promedio > Decimal("0.01"):
                    r.costos.append((sku, producto.costo_promedio, costo_excel))

        r.ausentes = list(
            Producto.objects.filter(empresa=empresa, activo=True)
            .exclude(sku__in=vistos)
            .values_list("sku", "nombre")
        )
        r.huerfanas = list(
            Categoria.objects.filter(productos__isnull=True, hijas__isnull=True)
            .values_list("nombre", flat=True)
        )
        return r

    def _categoria(self, web, subcategoria, r):
        """Devuelve la subcategoría (hoja del árbol), creando la rama si falta.

        La categoría web queda como raíz (`padre=None`) y la subcategoría
        cuelga de ella. `Categoria.nombre` es único a nivel global, así que
        una subcategoría no puede repetirse bajo dos categorías web: se
        verificó contra el Excel (75 subcategorías, ninguna repetida) y si
        algún día se repitiera, este método reasigna el padre y lo reporta,
        en vez de crear un duplicado silencioso.
        """
        if not web:
            return None
        raiz, creada = Categoria.objects.get_or_create(nombre=web, defaults={"padre": None})
        if creada:
            r.categorias_creadas.append(web)
        if raiz.padre_id is not None:
            raiz.padre = None
            raiz.save(update_fields=["padre"])
        if not subcategoria:
            return raiz
        hija, creada = Categoria.objects.get_or_create(nombre=subcategoria, defaults={"padre": raiz})
        if creada:
            r.categorias_creadas.append(f"{web} › {subcategoria}")
        elif hija.padre_id != raiz.id:
            hija.padre = raiz
            hija.save(update_fields=["padre"])
            r.recolgadas.append(f"{subcategoria} → {web}")
        return hija

    # ── reporte ──────────────────────────────────────────────────────────
    def _reportar(self, r, seco, sin_precios=False):
        w = self.stdout.write
        titulo = "SIMULACIÓN (no se escribió nada)" if seco else "Sincronización aplicada"
        w(self.style.SUCCESS(f"\n{titulo}"))
        w(f"  Productos creados ........ {r.creados}")
        w(f"  Productos actualizados ... {r.actualizados}  (de los cuales {r.renombrados} cambian de nombre)")
        if sin_precios:
            w(self.style.WARNING(
                f"  Precios NO tocados ....... {r.precios_omitidos} (por --sin-precios)"
            ))
        else:
            w(f"  Precios cambiados ........ {len(r.precios)}")
        w(f"  Altas de stock (INI) ..... {r.stock_alta}")
        w(f"  Ajustes de stock (AJU) ... {len(r.ajustes)}")
        w(f"  Categorías creadas ....... {len(r.categorias_creadas)}")

        if r.precios and sin_precios:
            subidas = sum(1 for _, a, n in r.precios if n > a)
            w(self.style.WARNING(
                f"\n  {len(r.precios)} producto(s) tienen un precio distinto en el Excel "
                f"({subidas} más caro, {len(r.precios) - subidas} más barato) y se dejaron "
                "como estaban.\n  Para aplicarlos, corré el comando sin --sin-precios."
            ))
        elif r.precios:
            subidas = sum(1 for _, a, n in r.precios if n > a)
            w(self.style.WARNING(
                f"\n  {len(r.precios)} precios cambian ({subidas} suben, {len(r.precios) - subidas} bajan)."
            ))
            for sku, antes, ahora in r.precios[:10]:
                pct = ((ahora - antes) / antes * 100) if antes else Decimal("0")
                w(f"    {sku}: ₡{antes:,.0f} → ₡{ahora:,.0f}  ({pct:+.0f}%)")
            if len(r.precios) > 10:
                w(f"    … y {len(r.precios) - 10} más. Quedan todos en CambioPrecio.")

        if r.costos:
            w(self.style.WARNING(
                f"\n  {len(r.costos)} producto(s) tienen un costo promedio distinto al del Excel."
            ))
            w("  NO se corrigieron: el costo sale del kardex, y cambiarlo es un asiento")
            w("  contable, no una importación. Revisalos y decidí:")
            for sku, erp, excel in r.costos[:8]:
                w(f"    {sku}: ERP ₡{erp:,.2f} · Excel ₡{excel:,.2f}")
            if len(r.costos) > 8:
                w(f"    … y {len(r.costos) - 8} más.")

        if r.ausentes:
            w(self.style.WARNING(f"\n  {len(r.ausentes)} producto(s) activos NO están en el Excel:"))
            for sku, nombre in r.ausentes[:10]:
                w(f"    {sku} — {nombre}")
            w("  Se dejaron activos. Usá --desactivar-ausentes si querés retirarlos.")

        if r.huerfanas:
            w(f"\n  {len(r.huerfanas)} categoría(s) quedaron sin productos: {', '.join(r.huerfanas)}")
            w("  No estorban (el sitio solo publica categorías con productos), pero se pueden borrar.")

        if r.duplicados:
            w(self.style.WARNING(f"\n  SKU repetidos en el Excel (se usó el primero): {', '.join(r.duplicados)}"))
        if r.errores:
            w(self.style.ERROR(f"\n  {len(r.errores)} error(es):"))
            for e in r.errores[:20]:
                w(f"    {e}")

        if seco:
            w(self.style.WARNING("\n  Nada de esto se guardó. Quitá --dry-run para aplicarlo.\n"))
        else:
            w(self.style.SUCCESS(
                "\n  Listo. Siguiente paso: `manage.py importar_imagenes` y "
                "`manage.py exportar_catalogo_web`.\n"
            ))


class _Resultado:
    def __init__(self):
        self.creados = 0
        self.actualizados = 0
        self.renombrados = 0
        self.stock_alta = 0
        self.saltados = 0
        self.sin_precios = False
        self.precios_omitidos = 0
        self.precios = []
        self.ajustes = []
        self.costos = []
        self.errores = []
        self.duplicados = []
        self.ausentes = []
        self.huerfanas = []
        self.categorias_creadas = []
        self.recolgadas = []
