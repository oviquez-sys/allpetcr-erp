"""Carga inicial del inventario desde el Excel real.

Uso:
    python manage.py importar_inventario "data/INVENTARIO REAL AL 6-7-2026   2.0.xlsx"

- Crea (si no existen) la empresa ALLPETCR.COM (régimen simplificado),
  la sucursal Central y la bodega Principal.
- Agrupa las categorías del Excel en familias; la original se conserva
  en `categoria_original` para la depuración pendiente.
- Cada producto entra al kardex como CARGA_INICIAL vía el servicio de
  dominio: mismo camino que usará cualquier movimiento futuro.
- Idempotente: los SKU ya existentes se omiten (reporta cuántos).
"""
from decimal import Decimal

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from catalogo.models import Categoria, Producto
from core.models import Empresa, Sucursal
from inventario.models import Bodega
from inventario.services import registrar_movimiento

FAMILIAS = {
    "Accesorios para mascotas": "Accesorios",
    "Entrenamiento": "Accesorios",
    "Areneros e individuales": "Gatos: areneros y rascadores",
    "Rascadores y casas gatos": "Gatos: areneros y rascadores",
    "Arnes de gatos y conejos": "Collares, correas y arneses",
    "ARNES DE PERRO": "Collares, correas y arneses",
    "BOZALES PERRO": "Collares, correas y arneses",
    "COLLAR PERRO": "Collares, correas y arneses",
    "COLLAR Y CORREA DE PERRO": "Collares, correas y arneses",
    "CORREAS": "Collares, correas y arneses",
    "PAÑUELO Y COLLAR DE GATO": "Collares, correas y arneses",
    "FUNDAS PARA AUTO": "Transporte y paseo",
    "CORREAS DE AUTO": "Transporte y paseo",
    "MOCHILA PARA MASCOTAS": "Transporte y paseo",
    "CAMAS Y MANTAS": "Camas y descanso",
    "MANTAS REFRESCANTES": "Camas y descanso",
    "COLLAR ISABELINO": "Salud e higiene",
    "HIGIENE ANIMAL": "Salud e higiene",
    "QUITA PELUSAS": "Salud e higiene",
    "JUGUETES": "Juguetes y peluches",
    "JUGUETES DE TELA": "Juguetes y peluches",
    "JUGUETES ELECTRONICOS": "Juguetes y peluches",
    "JUGUETES PARA GATOS": "Juguetes y peluches",
    "JUGUETES PLASTICOS": "Juguetes y peluches",
    "PELUCHES": "Juguetes y peluches",
    "PALA PARA COMIDA, BARRIL Y ALIMENTADOR": "Comederos y bebederos",
    "TASA Y BEBEDERO METALICO": "Comederos y bebederos",
    "TASA Y BEBEDEROS PLASTICOS": "Comederos y bebederos",
    "ROPA PARA MASCOTAS": "Ropa",
}


class Command(BaseCommand):
    help = "Importa el inventario inicial desde el Excel real de ALLPETCR"

    def add_arguments(self, parser):
        parser.add_argument("ruta_excel")

    @transaction.atomic
    def handle(self, *args, **opts):
        try:
            wb = openpyxl.load_workbook(opts["ruta_excel"], data_only=True)
        except FileNotFoundError:
            raise CommandError(f"No se encontró el archivo: {opts['ruta_excel']}")
        if "Inventario Real" not in wb.sheetnames:
            raise CommandError("El Excel no tiene la hoja 'Inventario Real'.")
        ws = wb["Inventario Real"]

        empresa, _ = Empresa.objects.get_or_create(
            nombre="ALLPETCR.COM", defaults={"regimen": Empresa.Regimen.SIMPLIFICADO}
        )
        sucursal, _ = Sucursal.objects.get_or_create(empresa=empresa, nombre="Central")
        bodega, _ = Bodega.objects.get_or_create(sucursal=sucursal, nombre="Principal")

        creados = omitidos = 0
        for fila in ws.iter_rows(min_row=2, values_only=True):
            nombre = fila[2]
            if not nombre:
                continue
            sku = str(fila[1]).strip()
            if Producto.objects.filter(sku=sku).exists():
                omitidos += 1
                continue

            cat_original = str(fila[3]).strip() if fila[3] else ""
            familia, _ = Categoria.objects.get_or_create(
                nombre=FAMILIAS.get(cat_original, "Accesorios")
            )
            costo = Decimal(str(fila[11] or 0)).quantize(Decimal("0.01"))
            precio = Decimal(str(fila[12] or 0)).quantize(Decimal("0.01"))
            cantidad = Decimal(str(fila[8] or 0))

            producto = Producto.objects.create(
                empresa=empresa,
                sku=sku,
                nombre=str(nombre).strip(),
                categoria=familia,
                categoria_original=cat_original,
                codigo_barras=str(fila[4]).strip() if fila[4] else "",
                presentacion=str(fila[5]).strip() if fila[5] else "",
                precio_venta=precio,
            )
            if cantidad > 0:
                registrar_movimiento(
                    producto=producto,
                    bodega=bodega,
                    tipo="INI",
                    cantidad=cantidad,
                    costo_unitario=costo,
                    referencia="CARGA-INICIAL 6-7-2026",
                    motivo="Importación del Excel de inventario real",
                )
            creados += 1

        self.stdout.write(self.style.SUCCESS(
            f"Importación completa: {creados} productos creados, {omitidos} omitidos (SKU ya existente)."
        ))
