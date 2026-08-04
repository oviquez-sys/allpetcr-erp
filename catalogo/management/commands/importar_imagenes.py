"""Extrae las fotos embebidas en el Excel y las asocia a cada producto.

Uso:
    python manage.py importar_imagenes                  # usa data/INVENTARIO_ALLPETCR.xlsx
    python manage.py importar_imagenes --archivo otro.xlsx --hoja "Otra hoja"

Cómo funciona
-------------
Las imágenes de un .xlsx viven en `xl/media/` y su posición está en
`xl/drawings/drawing1.xml`: cada una se ancla a una fila. Como la fila de la
foto es la fila del producto, se mapea fila → SKU leyendo la columna B de la
hoja. Es un mapeo por posición, no por nombre de archivo: los archivos de
`xl/media/` se llaman `image1.jpeg`, `image2.jpeg`… y no dicen a qué producto
pertenecen.

Por qué la hoja se pasa por nombre y no se adivina
--------------------------------------------------
Antes, si la hoja esperada no estaba, caía en `wb.active` — la hoja que quedó
seleccionada al guardar el archivo. Con el Excel del 02/08/2026 eso habría
leído "Supuestos" (25 filas de costos fijos) y mapeado imágenes contra basura
sin que nada fallara: cada producto habría quedado con la foto de otro. Ahora
aborta con un mensaje claro.
"""
import re
import zipfile
from pathlib import Path

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from catalogo.models import Producto

HOJA_POR_DEFECTO = "Inventario Claude"
ARCHIVO_POR_DEFECTO = "data/INVENTARIO_ALLPETCR.xlsx"
COL_SKU = 1  # columna B, 0-based


class Command(BaseCommand):
    help = "Extrae las imágenes del Excel y las asigna a cada producto por su fila"

    def add_arguments(self, parser):
        parser.add_argument("ruta_excel", nargs="?", default=ARCHIVO_POR_DEFECTO)
        parser.add_argument("--archivo", dest="archivo", default=None,
                            help="Alias de ruta_excel, por simetría con sincronizar_inventario.")
        parser.add_argument("--hoja", default=HOJA_POR_DEFECTO)

    @transaction.atomic
    def handle(self, *args, **opts):
        ruta = opts["archivo"] or opts["ruta_excel"]
        hoja = opts["hoja"]
        try:
            zf = zipfile.ZipFile(ruta)
        except (FileNotFoundError, zipfile.BadZipFile):
            raise CommandError(f"No se pudo abrir el Excel: {ruta}")

        with zf:
            nombres = zf.namelist()
            draw = next((n for n in nombres if re.match(r"xl/drawings/drawing\d+\.xml$", n)), None)
            if not draw:
                raise CommandError("El Excel no tiene imágenes embebidas (sin drawings).")
            xml = zf.read(draw).decode("utf-8", "ignore")
            rels = zf.read("xl/drawings/_rels/" + draw.split("/")[-1] + ".rels").decode("utf-8")
            rid_target = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels))

            # (fila 0-based) -> media file, tomando la primera imagen por fila
            fila_media = {}
            anchors = re.findall(
                r"<xdr:(?:twoCellAnchor|oneCellAnchor)\b.*?</xdr:(?:twoCellAnchor|oneCellAnchor)>",
                xml, re.S,
            )
            for a in anchors:
                mr = re.search(r"<xdr:from>.*?<xdr:row>(\d+)</xdr:row>", a, re.S)
                me = re.search(r'r:embed="(rId\d+)"', a)
                if not (mr and me):
                    continue
                fila = int(mr.group(1))
                target = rid_target.get(me.group(1), "")
                media = "xl/" + target.replace("../", "")
                fila_media.setdefault(fila, media)

            # Mapa fila 0-based del Excel -> SKU (columna B), leyendo la hoja.
            wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
            if hoja not in wb.sheetnames:
                disponibles = ", ".join(wb.sheetnames)
                wb.close()
                raise CommandError(
                    f"El Excel no tiene la hoja '{hoja}'. Hojas disponibles: {disponibles}.\n"
                    "  Pasá la correcta con --hoja. No se adivina: mapear imágenes contra "
                    "la hoja equivocada le pone a cada producto la foto de otro."
                )
            ws = wb[hoja]
            fila_sku = {}
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                # i es 0-based y coincide con la fila 0-based de los anchors
                sku = row[COL_SKU] if len(row) > COL_SKU else None
                if sku:
                    fila_sku[i] = str(sku).strip()
            wb.close()

            por_sku = {p.sku: p for p in Producto.objects.all()}
            destino = Path(settings.MEDIA_ROOT) / "productos"
            destino.mkdir(parents=True, exist_ok=True)

            guardadas = sin_producto = 0
            con_foto = set()
            for fila, media in fila_media.items():
                sku = fila_sku.get(fila)
                prod = por_sku.get(sku) if sku else None
                if prod is None:
                    sin_producto += 1
                    continue
                try:
                    data = zf.read(media)
                except KeyError:
                    continue
                ext = Path(media).suffix.lower() or ".png"
                nombre_archivo = f"{prod.sku}{ext}"
                (destino / nombre_archivo).write_bytes(data)
                rel = f"productos/{nombre_archivo}"
                if prod.imagen != rel:
                    prod.imagen = rel
                    prod.save(update_fields=["imagen"])
                guardadas += 1
                con_foto.add(prod.sku)

        self.stdout.write(self.style.SUCCESS(
            f"Listo: {guardadas} imágenes extraídas y asignadas a productos."
            + (f" ({sin_producto} imágenes sin producto coincidente)" if sin_producto else "")
        ))

        # Un producto sin foto se publica con el marcador "Foto pendiente" del
        # sitio: no se ve roto, pero conviene saber cuáles son para fotografiarlos.
        faltan = [
            (p.sku, p.nombre)
            for p in Producto.objects.filter(activo=True).order_by("nombre")
            if p.sku not in con_foto and not p.imagen
        ]
        if faltan:
            self.stdout.write(self.style.WARNING(
                f"\n  {len(faltan)} producto(s) activos quedaron sin foto:"
            ))
            for sku, nombre in faltan[:15]:
                self.stdout.write(f"    {sku} — {nombre}")
            if len(faltan) > 15:
                self.stdout.write(f"    … y {len(faltan) - 15} más.")
